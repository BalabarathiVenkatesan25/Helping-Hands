import io
from django.http import HttpResponse
from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from accounts.models import User, DonorProfile, OrphanageProfile
from donations.models import DonationRequest, Donation

def generate_pdf_report(report_type):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=colors.HexColor('#4F46E5'),
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'DocSub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#475569'),
        spaceAfter=20
    )
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11
    )
    header_style = ParagraphStyle(
        'HeaderText',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white
    )

    if report_type == 'donations':
        title = "Helping Hands - Donations Report"
        subtitle = "Detailed system-wide history of items and financial donations."
        headers = ["ID", "Donor", "Orphanage/Request", "Type", "Quantity/Amt", "Status", "Date"]
        
        donations = Donation.objects.all().order_by('-created_at')
        data = [[Paragraph(h, header_style) for h in headers]]
        
        for d in donations:
            target = d.request.title if d.request else "General Donation"
            val = f"${d.amount}" if d.donation_type == 'Money' else f"{d.quantity_donated} Items"
            
            row = [
                Paragraph(str(d.id), cell_style),
                Paragraph(d.donor.full_name, cell_style),
                Paragraph(target, cell_style),
                Paragraph(d.donation_type, cell_style),
                Paragraph(val, cell_style),
                Paragraph(d.status, cell_style),
                Paragraph(d.created_at.strftime('%Y-%m-%d'), cell_style)
            ]
            data.append(row)
            
        col_widths = [30, 100, 150, 50, 70, 70, 70]
        
    elif report_type == 'users':
        title = "Helping Hands - User Accounts Report"
        subtitle = "Directory of all registered donors, orphanages, and administrators."
        headers = ["ID", "Username", "Email", "Role", "Status", "Joined Date"]
        
        users = User.objects.all().order_by('-date_joined')
        data = [[Paragraph(h, header_style) for h in headers]]
        
        for u in users:
            role = "Admin"
            if u.is_donor:
                role = "Donor"
            elif u.is_orphanage:
                role = "Orphanage"
                
            status = "Active" if u.is_active else "Suspended"
            
            row = [
                Paragraph(str(u.id), cell_style),
                Paragraph(u.username, cell_style),
                Paragraph(u.email, cell_style),
                Paragraph(role, cell_style),
                Paragraph(status, cell_style),
                Paragraph(u.date_joined.strftime('%Y-%m-%d'), cell_style)
            ]
            data.append(row)
            
        col_widths = [30, 100, 150, 80, 80, 100]
        
    elif report_type == 'requests':
        title = "Helping Hands - Donation Requests Report"
        subtitle = "Tracking orphanage requirements and fulfillment metrics."
        headers = ["ID", "Orphanage", "Title", "Category", "Needed/Fulfilled", "Status", "Deadline"]
        
        requests = DonationRequest.objects.all().order_by('-created_at')
        data = [[Paragraph(h, header_style) for h in headers]]
        
        for r in requests:
            qty = f"{r.quantity_needed} / {r.quantity_fulfilled}"
            row = [
                Paragraph(str(r.id), cell_style),
                Paragraph(r.orphanage.orphanage_name, cell_style),
                Paragraph(r.title, cell_style),
                Paragraph(r.category.name if r.category else "Uncategorized", cell_style),
                Paragraph(qty, cell_style),
                Paragraph(r.status, cell_style),
                Paragraph(r.deadline.strftime('%Y-%m-%d'), cell_style)
            ]
            data.append(row)
            
        col_widths = [30, 110, 140, 70, 80, 50, 60]
    else:
        return None

    # Construct Document
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(subtitle, subtitle_style))
    story.append(Spacer(1, 15))
    
    # Table layout
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4F46E5')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
        ('TOPPADDING', (0,1), (-1,-1), 6),
    ]))
    story.append(table)
    
    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf

def generate_excel_report(report_type):
    wb = Workbook()
    ws = wb.active
    ws.title = report_type.capitalize()
    
    # Colors and styles
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=16, bold=True, color="4F46E5")
    
    if report_type == 'donations':
        ws.append(["Helping Hands - Donations Report"])
        ws.cell(row=1, column=1).font = title_font
        ws.append([]) # empty spacer
        
        headers = ["Donation ID", "Donor Name", "Recipient Request", "Type", "Quantity Donated", "Amount", "Status", "Date"]
        ws.append(headers)
        
        donations = Donation.objects.all().order_by('-created_at')
        for d in donations:
            target = d.request.title if d.request else "General Donation"
            ws.append([
                d.id,
                d.donor.full_name,
                target,
                d.donation_type,
                d.quantity_donated if d.donation_type == 'Items' else '',
                d.amount if d.donation_type == 'Money' else '',
                d.status,
                d.created_at.strftime('%Y-%m-%d')
            ])
            
    elif report_type == 'users':
        ws.append(["Helping Hands - User Accounts Report"])
        ws.cell(row=1, column=1).font = title_font
        ws.append([]) # empty spacer
        
        headers = ["User ID", "Username", "Email", "Role", "Status", "Date Joined"]
        ws.append(headers)
        
        users = User.objects.all().order_by('-date_joined')
        for u in users:
            role = "Admin"
            if u.is_donor:
                role = "Donor"
            elif u.is_orphanage:
                role = "Orphanage"
            status = "Active" if u.is_active else "Suspended"
            ws.append([
                u.id,
                u.username,
                u.email,
                role,
                status,
                u.date_joined.strftime('%Y-%m-%d')
            ])
            
    elif report_type == 'requests':
        ws.append(["Helping Hands - Donation Requests Report"])
        ws.cell(row=1, column=1).font = title_font
        ws.append([]) # empty spacer
        
        headers = ["Request ID", "Orphanage Name", "Request Title", "Category", "Quantity Needed", "Quantity Fulfilled", "Priority", "Status", "Deadline"]
        ws.append(headers)
        
        requests = DonationRequest.objects.all().order_by('-created_at')
        for r in requests:
            ws.append([
                r.id,
                r.orphanage.orphanage_name,
                r.title,
                r.category.name if r.category else "Uncategorized",
                r.quantity_needed,
                r.quantity_fulfilled,
                r.priority,
                r.status,
                r.deadline.strftime('%Y-%m-%d')
            ])
            
    # Format Headers (Row 3 is header)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        # Skip rows 1 and 2 for length checks (since title might span far)
        for cell in col[2:]:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    excel_data = buffer.getvalue()
    buffer.close()
    return excel_data
